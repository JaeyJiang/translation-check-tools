# encoding=utf-8
import sys
import time
import threading

import openpyxl
import xlrd

"""
@author: v_jjyjiang 
@Description: TODO(这里用一句话描述这个模块的作用)
@contact: jaey_summer@qq.com
@software: PyCharm
@file: excel_utils.py
"""


class ExcelUtils(object):
    def __init__(self):
        pass

    @staticmethod
    def cvt_xls_to_xlsx(one_file_path, dst_file_path):
        book_xls = xlrd.open_workbook(one_file_path)
        book_xlsx = openpyxl.Workbook()

        sheet_names = book_xls.sheet_names()
        for sheet_index in range(0, len(sheet_names)):
            sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
            # if sheet_index == 0:
            #     sheet_xlsx = book_xlsx.get_sheet_by_name("Sheet1")
            #     sheet_xlsx.title = sheet_names[sheet_index]
            # else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])
            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row,
                                                                                              col)
        book_xlsx.remove_sheet(book_xlsx.get_sheet_by_name("Sheet"))
        book_xlsx.save(dst_file_path)
